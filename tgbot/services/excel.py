from datetime import datetime
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
import os


class ExcelFile:
    def __init__(self):
        self.tickers_path = f"{os.getcwd()}/tickers_list.xlsx"
        self.result_path = f"{os.getcwd()}/result_list.xlsx"

    @staticmethod
    def __create_file(path: str, head: tuple, data: List[tuple]):
        wb = Workbook()
        ws = wb.active
        ws.append(head)
        title_ft = Font(bold=True)
        for row in ws['A1:T1']:
            for cell in row:
                cell.font = title_ft
        for item in data:
            ws.append(item)
        wb.save(path)

    def create_tickers_file(self, tickers: list):
        tickers = [(i,) for i in tickers]
        self.__create_file(path=self.tickers_path, head=("Тикер",), data=tickers)

    def read_tickers_file(self) -> Optional[list]:
        wb = load_workbook(filename=self.tickers_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2):
            try:
                if not row[0].value:
                    continue
                data.append(row[0].value)
            except (ValueError, Exception):
                return None
        return data

    def create_stocks_file(self, stocks: List[dict]):
        head = (
                "Тикер",
                "Название компании",
                "P/S (Price to Sales)",
                "Рост выручки (Total Revenue Growth Rate (TTM)",
                "Рост выручки / P/S",
                "Кол-во лет"
            )
        data = []
        for item in stocks:
            data.append(
                (
                    item["ticker"],
                    item["name"],
                    item["price_to_sales"],
                    item["total_revenue_ttm"],
                    item["tr_div_ps"],
                    item["years"],
                )
            )
        self.__create_file(path=self.result_path, head=head, data=data)
